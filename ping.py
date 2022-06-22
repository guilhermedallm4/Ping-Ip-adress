import socket
import openpyxl 
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook


##criar um excel(book)

book = openpyxl.Workbook()

remove  = book['Sheet']
book.remove_sheet(remove)
#Criando pagina
book.create_sheet('Sincronizacao')
#Como selecionar pagina
sinc_page =  book['Sincronizacao']
sinc_page.append(['Empresa','Ip', 'Porta', 'Valor'])

for col_range in range(1, 5):
    cell_title = sinc_page.cell(1, col_range)
    cell_title.fill = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")


def test_port(ip, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = sock.connect_ex((ip, port))
    sock.close()

    if result == 0:
        return 'True'
    else:
        return 'False'

global list_ip 
global list_port 
global list_name
global carac
carac = '\n'
list_name = []
list_ip = []
list_port = []


with open('name.txt', 'r') as client:
    for name_client in client:
        list_name.append(str(name_client).replace(carac, ''))    


with open('hosts.txt', 'r') as archive:
    for ip_adress in archive:
        list_ip.append(str(ip_adress).replace(carac, ''))

with open('ports.txt', 'r') as port:
        for port_adress in port:
            list_port.append(int(str(port_adress).replace(carac, '')))

        
'''print(list_name)    
print(list_ip)
print(list_port)'''
global auxInt
for i in range((len(list_ip))):
    status = test_port(list_ip[i], list_port[i])
    print(status)
    print(i+1)
    sinc_page.append([list_name[i], list_ip[i], str(list_port[i]), status])
    aux = str(i)
    auxInt = i + 2
    if status == 'True':
        for col_range in range(1, 5):
            cell_title = sinc_page.cell(auxInt, col_range)
            cell_title.fill = PatternFill(start_color="0000FF00", end_color="0000FF00", fill_type="solid")
    else:
       for col_range in range(1, 5):
            cell_title = sinc_page.cell(auxInt, col_range)
            cell_title.fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")


book.save('Sincronizacao.xlsx')